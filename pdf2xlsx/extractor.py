"""Column extraction from converted Excel workbooks.

Lets the user pull a standard set of columns (Material Code, Item
Description, EAN No., Quantity, Unit Base Cost) out of the converted Excel
files and consolidate them into a single workbook. The selected columns are
matched against the table headers with a tolerant, synonym-based match, so
common naming variations ("Qty", "EAN Number", "Unit Cost", …) are found.
"""

import os
import re
import threading

from openpyxl import Workbook, load_workbook

from . import excelio as x
from .utils import unique_path

FIELD_LABELS = [
    ("material_code", "Material Code"),
    ("item_description", "Item Description"),
    ("ean", "EAN No."),
    ("quantity", "Quantity"),
    ("unit_cost", "Unit Base Cost"),
]
FIELD_LABEL_MAP = dict(FIELD_LABELS)

SYNONYMS = {
    "material_code": ["material code", "material", "material no",
                      "item code", "item no", "product code",
                      "article no", "artikel code", "sku", "ref no",
                      "reference", "matcode", "code of material"],
    "item_description": ["item description", "item desc", "description",
                         "description of goods", "description of item",
                         "description of material", "product description",
                         "desc", "item"],
    "ean": ["ean", "ean no", "ean number", "ean-13", "ean13", "barcode",
            "bar code", "gtin", "gtin no", "upc", "upc no", "barcode ean"],
    "quantity": ["quantity", "qty", "qty no", "qty pcs", "qty(pcs)",
                 "no of pcs", "no of pieces", "number of pieces", "pieces",
                 "quantity of items", "total qty"],
    "unit_cost": ["unit base cost", "unit cost", "base cost",
                  "unit base price", "unit price", "base price",
                  "cost", "price", "unit base"],
}

HEADER_LIMIT = 60  # how many rows are scanned looking for a header
STRING_FIELDS = {"ean", "material_code"}  # keep as text to avoid precision loss


class ExtractionCancelled(Exception):
    """Raised when the user cancels a running extraction."""


def _as_string(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _norm(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()


def _matches(cell_text, field_key):
    """Tolerant header-cell match: exact first, then one-way containment."""
    n = _norm(cell_text)
    if not n:
        return False
    for syn in SYNONYMS.get(field_key, []):
        s = _norm(syn)
        if s == n:
            return True
        if len(s) >= 4 and (s in n or n in s):
            return True
    return False


def _find_header(ws, fields):
    """Locate the table header row and map fields to column indexes.

    Returns ``(header_row, col_map)`` where ``col_map`` is
    ``{field_key: zero_based_column}``. ``header_row`` is None when no
    header with any requested field is found on the sheet.
    """
    best_row = None
    best_score = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True, max_row=HEADER_LIMIT), start=1):
        score = 0
        for f in fields:
            if any(v is not None and _matches(v, f) for v in row):
                score += 1
        if score > best_score:
            best_score = score
            best_row = idx
        if best_score == len(fields):
            break
    if best_row is None or best_score == 0:
        return None, {}

    header_values = next(ws.iter_rows(values_only=True,
                                      min_row=best_row, max_row=best_row))
    col_map = {}
    for f in fields:
        for ci, v in enumerate(header_values):
            if v is not None and _matches(v, f):
                col_map[f] = ci
                break
    return best_row, col_map


def _extract_file(path, fields):
    """Return ``(rows, notes)`` for one workbook, scanning every sheet."""
    wb = load_workbook(path, read_only=True)
    rows_out = []
    notes = []
    try:
        matches = []
        for ws in wb.worksheets:
            header_row, col_map = _find_header(ws, fields)
            if header_row is not None:
                matches.append((ws, header_row, col_map))
        if not matches:
            return [], []

        # Prefer dedicated table sheets (converter names them "… Table N");
        # a prose sheet may name the columns in a sentence, so ignore non-table
        # sheets whenever at least one real table sheet matched.
        table_matches = [m for m in matches if "table" in m[0].title.lower()]
        if table_matches:
            matches = table_matches

        for ws, header_row, col_map in matches:
            missing = [f for f in fields if f not in col_map]
            for f in missing:
                notes.append(
                    f"'{os.path.basename(path)}' → '{ws.title}': "
                    f"column '{FIELD_LABEL_MAP[f]}' not found")
            for row in ws.iter_rows(values_only=True, min_row=header_row + 1):
                if all(v is None or str(v).strip() == "" for v in row):
                    break  # end of the table's data block
                record = {}
                for f, ci in col_map.items():
                    value = row[ci] if ci < len(row) else None
                    if f in STRING_FIELDS and value is not None:
                        value = _as_string(value)
                    record[f] = value
                rows_out.append(record)
    finally:
        wb.close()
    return rows_out, notes


def extract_fields(input_paths, out_path, fields,
                   cancel_event=None, progress_cb=None):
    """Consolidate the requested ``fields`` from every workbook into one xlsx.

    Returns ``{path, rows, columns, notes}``. ``fields`` is an ordered list
    of keys from ``FIELD_LABELS``.
    """
    if not fields:
        raise ValueError("Select at least one field to extract.")
    out_path = unique_path(out_path)
    labels = [FIELD_LABEL_MAP[f] for f in fields]

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Extracted")
    for ci, label in enumerate(labels, start=1):
        x.write_cell(ws, 1, ci, label, header=True)
    ws.freeze_panes = "A2"

    total = 0
    notes = []
    for i, path in enumerate(input_paths):
        if cancel_event is not None and cancel_event.is_set():
            raise ExtractionCancelled()
        if progress_cb:
            progress_cb(i + 1, len(input_paths), os.path.basename(path))
        try:
            rows_out, file_notes = _extract_file(path, fields)
        except Exception as exc:
            notes.append(f"{os.path.basename(path)}: could not be read ({exc})")
            continue
        if not rows_out:
            notes.append(f"{os.path.basename(path)}: no matching table found")
        for record in rows_out:
            ws.append([record.get(f) for f in fields])
        total += len(rows_out)
        notes.extend(file_notes)

    x.autofit_columns(ws)

    part = out_path + ".part"
    try:
        wb.save(part)
        os.replace(part, out_path)
    finally:
        wb.close()
        if os.path.exists(part):
            try:
                os.remove(part)
            except OSError:
                pass

    return {"path": out_path, "rows": total, "columns": labels, "notes": notes}
