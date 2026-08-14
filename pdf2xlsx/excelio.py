"""openpyxl helpers: sheet creation, cell styling, sheet copying for merges."""

import io
import os
import re
import tempfile
from copy import copy

from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INT_RE = re.compile(r"^[+-]?\d+$")
FLOAT_RE = re.compile(r"^[+-]?\d+\.\d+(?:[eE][+-]?\d+)?$")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
THIN = Side(style="thin", color="8A8F98")
BODY_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def try_numeric(value):
    """Convert plain numeric strings to int/float so cells are usable in Excel.
    Values with leading zeros, separators, or non-numeric content stay strings
    (nothing is lost)."""
    if not isinstance(value, str):
        return value
    s = value.strip()
    if s == "":
        return None
    if INT_RE.match(s):
        return int(s)
    if FLOAT_RE.match(s):
        return float(s)
    return value


def sanitize_sheet_title(title):
    """Excel forbids ``[]:*?/\\`` and limits titles to 31 chars."""
    for ch in r"[]:*?/\\":
        title = title.replace(ch, "_")
    return (title.strip() or "Sheet")[:31]


def unique_sheet_title(workbook, title):
    """Return a sheet title not already used in ``workbook``."""
    title = sanitize_sheet_title(title)
    base = title
    n = 1
    while title in workbook.sheetnames:
        suffix = f" ({n})"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    return title


def prefixed_sheet_title(workbook, prefix, title):
    """Build ``prefix – title`` keeping the full sheet ``title`` and
    truncating the prefix when the combined name would exceed 31 chars."""
    sep = " – "
    combined = sanitize_sheet_title(f"{prefix}{sep}{title}")
    if len(combined) <= 31:
        return unique_sheet_title(workbook, combined)
    budget = 31 - len(sanitize_sheet_title(title))
    if budget > len(sep):
        p = sanitize_sheet_title(prefix[: budget - len(sep)])
        return unique_sheet_title(workbook, f"{p}{sep}{title}")
    return unique_sheet_title(workbook, title)


def write_cell(ws, row, col, value, *, header=False, bold=False, wrap=True):
    """Write ``value`` into a cell applying minimal consistent styling."""
    cell = ws.cell(row=row, column=col, value=try_numeric(value))
    if header:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    elif bold:
        cell.font = Font(bold=True, size=10)
    else:
        cell.font = BODY_FONT
    if wrap and isinstance(value, str) and "\n" in value:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = BODY_BORDER
    return cell


def autofit_columns(ws, min_width=8, max_width=60):
    """Approximate column widths from the longest visible cell."""
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        width = min_width
        for cell in col_cells:
            if cell.value is None:
                continue
            lines = str(cell.value).split("\n")
            longest = max(len(line) for line in lines)
            width = max(width, min(longest + 2, max_width))
        ws.column_dimensions[letter].width = width


def autofit_row_heights(ws):
    """Give wrapped rows a height matching their line count."""
    for row_cells in ws.iter_rows():
        lines = 1
        for cell in row_cells:
            if isinstance(cell.value, str):
                lines = max(lines, cell.value.count("\n") + 1)
        ws.row_dimensions[row_cells[0].row].height = max(15.0, lines * 15.0)


# --------------------------------------------------------------------------
# Sheet copying (used by the merger so formatting/data are never dropped)
# --------------------------------------------------------------------------


def copy_style(src_cell, dst_cell):
    """Copy font, fill, border, alignment, number format and protection."""
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)


def _anchor_coords(anchor):
    """Return (col_index, row_index) for an openpyxl image anchor."""
    if isinstance(anchor, OneCellAnchor):
        return anchor._from.col, anchor._from.row  # noqa: SLF001
    if isinstance(anchor, str):
        col_str = "".join(ch for ch in anchor if ch.isalpha())
        row_str = "".join(ch for ch in anchor if ch.isdigit())
        if not col_str or not row_str:
            return None
        from openpyxl.utils import column_index_from_string

        return column_index_from_string(col_str) - 1, int(row_str) - 1
    return None


def copy_images(src_ws, dst_ws, row_offset=0, col_offset=0):
    """Re-embed every picture from ``src_ws`` into ``dst_ws``."""
    copied = 0
    for img in list(getattr(src_ws, "_images", [])):
        try:
            data = img._data()  # noqa: SLF001
            ext = (img.format or "png").lower().lstrip(".")
            if ext not in ("png", "jpg", "jpeg", "gif", "bmp", "tiff", "tif"):
                ext = "png"
            with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as fh:
                fh.write(data)
                tmp_path = fh.name
            new_img = XLImage(tmp_path)
            new_img.width = int(getattr(img, "width", 0)) or new_img.width
            new_img.height = int(getattr(img, "height", 0)) or new_img.height
            anchor = getattr(img, "anchor", None)
            coords = _anchor_coords(anchor)
            if coords is not None:
                col, row = coords
                marker = AnchorMarker(col=col + col_offset, colOff=0,
                                      row=row + row_offset, rowOff=0)
                size = XDRPositiveSize2D(cx=int(new_img.width * 9525),
                                         cy=int(new_img.height * 9525))
                new_img.anchor = OneCellAnchor(_from=marker, ext=size)
            else:
                new_img.anchor = anchor
            dst_ws.add_image(new_img)
            copied += 1
        except Exception:
            continue
    return copied


def add_image_dicts(dst_ws, images, row_offset=0, col_offset=0):
    """Re-embed pre-extracted image records (see merger._extract_images)."""
    added = 0
    for rec in images:
        try:
            ext = (rec.get("ext") or "png").lower().lstrip(".")
            data = rec.get("data")
            if not data:
                continue
            new_img = XLImage(io.BytesIO(data))
            new_img.width = rec.get("cx") or new_img.width
            new_img.height = rec.get("cy") or new_img.height
            marker = AnchorMarker(col=rec["col"] + col_offset, colOff=0,
                                  row=rec["row"] + row_offset, rowOff=0)
            size = XDRPositiveSize2D(cx=int(new_img.width * 9525),
                                     cy=int(new_img.height * 9525))
            new_img.anchor = OneCellAnchor(_from=marker, ext=size)
            dst_ws.add_image(new_img)
            added += 1
        except Exception:
            continue
    return added


def copy_sheet_full(src_ws, dst_ws, images=None):
    """Copy all cells, styles, dimensions, merges, panes and images.

    ``images`` is an optional list of pre-extracted image records (the
    merger provides these because openpyxl cannot read pictures back from
    an existing workbook). When omitted, ``src_ws._images`` is used."""
    for col_letter, dim in src_ws.column_dimensions.items():
        target = dst_ws.column_dimensions[col_letter]
        target.width = dim.width
        target.hidden = dim.hidden
    for row_idx, dim in src_ws.row_dimensions.items():
        target = dst_ws.row_dimensions[row_idx]
        target.height = dim.height
        target.hidden = dim.hidden
    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))
    for row in src_ws.iter_rows():
        for cell in row:
            target = dst_ws.cell(row=cell.row, column=cell.column)
            target.value = cell.value
            if cell.has_style:
                copy_style(cell, target)
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes
    if src_ws.auto_filter.ref:
        dst_ws.auto_filter.ref = src_ws.auto_filter.ref
    if images is not None:
        add_image_dicts(dst_ws, images)
    else:
        copy_images(src_ws, dst_ws)


def append_sheet_data(src_ws, dst_ws, images=None):
    """Vertically append ``src_ws`` data (from row 2) below ``dst_ws`` data.
    Skips a repeated header row and preserves styles and images."""
    src_max_row = src_ws.max_row
    src_max_col = src_ws.max_column
    if src_max_row <= 1 or src_max_col <= 0:
        return 0
    offset = dst_ws.max_row
    start_src = 2
    header_values = [str(c.value).strip() for c in src_ws[1]]
    dst_header_values = [str(c.value).strip() for c in dst_ws[1]] if dst_ws.max_row >= 1 else []
    if header_values and header_values != dst_header_values:
        start_src = 1
    for row in src_ws.iter_rows(min_row=start_src, max_row=src_max_row):
        for cell in row:
            if cell.value is None:
                continue
            target = dst_ws.cell(row=cell.row - start_src + 1 + offset,
                                 column=cell.column)
            target.value = cell.value
            if cell.has_style:
                copy_style(cell, target)
    for merged in src_ws.merged_cells.ranges:
        mr = (merged.min_row - start_src + 1 + offset,
              merged.min_col,
              merged.max_row - start_src + 1 + offset,
              merged.max_col)
        if mr[2] >= mr[0] and mr[3] >= mr[1]:
            try:
                dst_ws.merge_cells(start_row=mr[0], start_column=mr[1],
                                   end_row=mr[2], end_column=mr[3])
            except Exception:
                pass
    if images is not None:
        add_image_dicts(dst_ws, images, row_offset=offset)
    else:
        copy_images(src_ws, dst_ws, row_offset=offset)
    return src_max_row - start_src + 1


def rows_values(ws):
    """Return the sheet as a list of row-lists of raw values."""
    return [[c.value for c in row] for row in ws.iter_rows()]
