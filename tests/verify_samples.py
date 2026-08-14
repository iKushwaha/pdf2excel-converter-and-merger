#!/usr/bin/env python3
"""Headless end-to-end verification of the converter and merger.

Converts the sample PDFs, inspects the generated workbooks, then merges two
of them in both modes and checks the results. Run from the project root:
    .venv/bin/python tests/verify_samples.py
"""

import os
import shutil
import sys
import zipfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook  # noqa: E402

from pdf2xlsx import converter, extractor, merger  # noqa: E402

HERE = os.path.dirname(__file__)
SAMPLES = os.path.join(HERE, "samples")
OUT = os.path.join(HERE, "output")

FAILURES = []
CHECKS = []


def check(name, condition, detail=""):
    status = "PASS" if condition else "FAIL"
    CHECKS.append((name, condition))
    print(f"  [{status}] {name}" + (f" — {detail}" if detail and not condition else ""))


def collect_sheet_values(ws):
    return {str(c.value) for row in ws.iter_rows() for c in row if c.value is not None}


def media_count(path):
    with zipfile.ZipFile(path) as z:
        return len([n for n in z.namelist() if n.startswith("xl/media/")])


def wipe_output():
    """Remove generated files from the output dir. Some filesystems create
    AppleDouble (._*) metadata files on write; delete files, not the dir."""
    if not os.path.isdir(OUT):
        return
    for name in os.listdir(OUT):
        if name.startswith("._"):
            continue  # filesystem metadata, harmless
        try:
            os.remove(os.path.join(OUT, name))
        except OSError:
            pass


def run():
    wipe_output()
    os.makedirs(OUT, exist_ok=True)

    pdfs = [p for p in os.listdir(SAMPLES)
            if p.endswith(".pdf") and not p.startswith("._")
            and p != "confidential.pdf"]
    print(f"\nFound {len(pdfs)} sample PDF(s)")

    results = {}
    for pdf in sorted(pdfs):
        path = os.path.join(SAMPLES, pdf)
        print(f"\nConverting {pdf} …")
        res = converter.convert_pdf_to_excel(path, OUT)
        results[pdf] = res
        print(f"  -> {os.path.basename(res['path'])}")
        print(f"     sheets: {res['sheets']}")
        if res["warnings"]:
            print(f"     warnings: {res['warnings']}")

    # ---------------- report_tables (tables + continuation) ----------------
    res = results["report_tables.pdf"]
    check("report: output exists", os.path.exists(res["path"]))
    wb = load_workbook(res["path"])
    sheets = wb.sheetnames
    check("report: has a Table worksheet", any("Table" in s for s in sheets),
          f"sheets={sheets}")
    table_sheets = [s for s in sheets if "Table" in s]
    check("report: table continued across pages (single sheet)",
          len(table_sheets) == 1, f"table sheets={table_sheets}")
    ws = wb[table_sheets[0]]
    vals = collect_sheet_values(ws)
    check("report: header preserved", "Item" in vals and "Amount" in vals)
    check("report: page-1 rows preserved", "Widget A" in vals and "Gadget C" in vals)
    check("report: continuation rows preserved", "Tool N" in vals and "Gadget Q" in vals)
    text_sheet = wb["Page 1 – Text & Images"]
    text_vals = collect_sheet_values(text_sheet)
    check("report: free text preserved",
          any("report summarises product sales" in str(v) for v in text_vals))
    check("report: page-2 text sheet exists", "Page 2 – Text & Images" in sheets)
    wb.close()

    # ---------------- notes_text (plain text) ------------------------------
    res = results["notes_text.pdf"]
    wb = load_workbook(res["path"])
    ws = wb["Page 1 – Text & Images"]
    vals = {str(c.value) for c in ws["A"] if c.value is not None}
    check("notes: paragraph text preserved",
          any("Meeting Notes" in v for v in vals))
    check("notes: multi-line action item preserved",
          any("Alice to share" in v for v in vals))
    check("notes: page 2 present", "Page 2 – Text & Images" in wb.sheetnames)
    wb.close()

    # ---------------- scan_image (image embedding) --------------------------
    res = results["scan_image.pdf"]
    wb = load_workbook(res["path"])
    ws = wb["Page 1 – Text & Images"]
    check("scan: image embedded", media_count(res["path"]) >= 1)
    vals = {str(c.value) for c in ws["A"] if c.value is not None}
    check("scan: text preserved", any("embedded as a picture" in v for v in vals))
    tbs = [s for s in wb.sheetnames if "Table" in s]
    check("scan: small table captured", len(tbs) == 1, f"tables={tbs}")
    wb.close()

    # ---------------- encrypted PDF ------------------------------------------
    print("\nEncrypted PDF handling …")
    enc_path = os.path.join(SAMPLES, "confidential.pdf")
    try:
        converter.convert_pdf_to_excel(enc_path, OUT)
        check("encrypted: raises EncryptedPDF without password", False)
    except converter.EncryptedPDF:
        check("encrypted: raises EncryptedPDF without password", True)
    except Exception as exc:
        check("encrypted: raises EncryptedPDF without password", False, str(exc))
    res = converter.convert_pdf_to_excel(enc_path, OUT, password="secret")
    wb = load_workbook(res["path"])
    ws = wb[wb.sheetnames[0]]
    vals = {str(c.value) for c in ws["A"] if c.value is not None}
    check("encrypted: converts with correct password",
          any("Confidential Monthly Summary" in v for v in vals))
    wb.close()

    # ---------------- unicode / spaces in filename ---------------------------
    print("\nUnicode filename handling …")
    src = os.path.join(SAMPLES, "notes_text.pdf")
    weird = os.path.join(SAMPLES, "München Bérlin report (final).pdf")
    shutil.copy2(src, weird)
    res = converter.convert_pdf_to_excel(weird, OUT)
    check("unicode: output created",
          os.path.exists(res["path"]), res["path"])
    check("unicode: base name preserved",
          os.path.basename(res["path"]).startswith("München Bérlin report (final)"),
          os.path.basename(res["path"]))
    os.remove(weird)

    # ---------------- merger -----------------------------------------------
    print("\nMerging report_tables + notes_text + scan_image (separate sheets) …")
    merged = merger.merge_excels(
        [results["report_tables.pdf"]["path"], results["notes_text.pdf"]["path"],
         results["scan_image.pdf"]["path"]],
        os.path.join(OUT, "Merged_test.xlsx"), mode="sheets")
    wb = load_workbook(merged["path"])
    check("merge-sheets: all source sheets present",
          len(wb.sheetnames) >= 7, f"sheets={wb.sheetnames}")
    check("merge-sheets: naming kept source prefix",
          any(s.startswith("report") for s in wb.sheetnames) and
          any(s.startswith("notes") for s in wb.sheetnames))
    check("merge-sheets: images preserved",
          media_count(merged["path"]) >= 1)
    wb.close()

    print("\nMerging report_tables + notes_text (append mode) …")
    merged2 = merger.merge_excels(
        [results["report_tables.pdf"]["path"], results["notes_text.pdf"]["path"]],
        os.path.join(OUT, "Merged_append_test.xlsx"), mode="append")
    wb = load_workbook(merged2["path"])
    append_sheets = [s for s in wb.sheetnames if "Table 1" in s]
    check("merge-append: matching tables appended into one sheet",
          len(append_sheets) == 1, f"sheets={wb.sheetnames}")
    ws = wb[append_sheets[0]] if append_sheets else wb[wb.sheetnames[0]]
    vals = collect_sheet_values(ws)
    check("merge-append: rows from both files present",
          "Widget A" in vals and "Gadget Q" in vals)
    wb.close()

    # ---------------- extraction of the five catalogue fields ---------------
    print("\nExtracting catalogue fields …")
    fields = ["material_code", "item_description", "ean", "quantity", "unit_cost"]
    extracted = extractor.extract_fields(
        [results["catalog.pdf"]["path"]],
        os.path.join(OUT, "Extracted_test.xlsx"), fields)
    check("extract: output exists", os.path.exists(extracted["path"]))
    check("extract: all four rows captured", extracted["rows"] == 4,
          f"rows={extracted['rows']}")
    check("extract: all five columns mapped", len(extracted["columns"]) == 5,
          f"columns={extracted['columns']}")
    wb = load_workbook(extracted["path"])
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    check("extract: headers written in order",
          header == ["Material Code", "Item Description", "EAN No.",
                     "Quantity", "Unit Base Cost"], f"header={header}")
    row2 = [c.value for c in ws[2]]
    check("extract: first data row correct",
          row2 == ["MC-1001", "Aluminium Screw 3mm", "4002355021235",
                   1500, 0.02], f"row2={row2}")
    eans = {c.value for r in ws.iter_rows(min_row=2) for c in r[:3]}
    check("extract: all EANs present",
          {"4002355021235", "4002355021242", "4002355021259",
           "4002355021266"}.issubset(eans))
    wb.close()

    print("\n" + "=" * 56)
    failed = [n for n, ok in CHECKS if not ok]
    print(f"Passed {len(CHECKS) - len(failed)}/{len(CHECKS)} checks.")
    if failed:
        print("FAILED:", ", ".join(failed))
        return 1
    print("All checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(run())
