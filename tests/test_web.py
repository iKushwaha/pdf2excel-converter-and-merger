#!/usr/bin/env python3
"""Headless tests for the PDF → Excel web app (Flask test client).

Run from the project root:
    .venv/bin/python tests/test_web.py
"""

import io
import os
import shutil
import sys
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import pymupdf as fitz  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from pdf2xlsx.web.server import create_app  # noqa: E402

FAILURES = []
CHECKS = []


def check(name, condition, detail=""):
    status = "PASS" if condition else "FAIL"
    CHECKS.append((name, condition))
    print(f"  [{status}] {name}" + (f" — {detail}" if detail and not condition else ""))


def make_sample_pdf(path):
    """Small table PDF with the extraction column names, plus an encrypted one."""
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 54), "Product Catalogue", fontsize=18)
    header = ["Material Code", "Item Description", "EAN No.",
              "Quantity", "Unit Base Cost"]
    rows = [
        ["MC-2001", "Hex Bolt 6mm", "4002355022001", "900", "0.05"],
        ["MC-2002", "Hex Nut 6mm", "4002355022002", "1500", "0.03"],
    ]
    x0, y0, x1, y1 = 72, 90, 520, 180
    data = [header] + rows
    nrows, ncols = len(data), len(data[0])
    step_x = (x1 - x0) / ncols
    step_y = (y1 - y0) / nrows
    for i in range(nrows + 1):
        y = y0 + i * step_y
        page.draw_line((x0, y), (x1, y), color=(0.1, 0.1, 0.1), width=0.7)
    for j in range(ncols + 1):
        x = x0 + j * step_x
        page.draw_line((x, y0), (x, y1), color=(0.1, 0.1, 0.1), width=0.7)
    for r, row in enumerate(data):
        for c, value in enumerate(row):
            page.insert_text((x0 + c * step_x + 4, y0 + r * step_y + step_y * 0.4),
                             str(value), fontsize=8)
    doc.save(path)
    doc.close()

    enc = path.replace(".pdf", "_enc.pdf")
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 80), "Confidential", fontsize=16)
    page.insert_text((72, 110), "Locked summary.", fontsize=10)
    doc.save(enc, encryption=fitz.PDF_ENCRYPT_AES_256,
             user_pw="secret", owner_pw="secret")
    doc.close()


def media_count(path):
    import zipfile

    with zipfile.ZipFile(path) as z:
        return len([n for n in z.namelist() if n.startswith("xl/media/")])


def main():
    tmp = tempfile.mkdtemp(prefix="pdf2xlsx-web-test-")
    data_dir = os.path.join(tmp, "data")
    sample = os.path.join(tmp, "catalog.pdf")
    try:
        make_sample_pdf(sample)

        app = create_app(data_dir=data_dir)
        app.config["SECRET_KEY"] = "test-secret"
        client = app.test_client()

        # ---- health -------------------------------------------------------
        r = client.get("/health")
        check("GET /health returns ok", r.status_code == 200 and r.get_json()["ok"])

        # ---- index --------------------------------------------------------
        r = client.get("/")
        check("GET / serves the web UI",
              r.status_code == 200 and b"PDF" in r.data and b"dropzone" in r.data)

        # ---- fields -------------------------------------------------------
        r = client.get("/api/fields")
        fields = r.get_json()["fields"]
        check("GET /api/fields lists extraction fields",
              r.status_code == 200 and {f["key"] for f in fields}
              >= {"material_code", "ean", "quantity"})

        # ---- upload + convert ----------------------------------------------
        with open(sample, "rb") as fh:
            r = client.post("/api/upload",
                            data={"pdfs": (fh, "catalog.pdf")},
                            content_type="multipart/form-data")
        payload = r.get_json()
        check("POST /api/upload converts the PDF",
              r.status_code == 200 and payload["ok"]
              and payload["files"][0]["ok"])
        xlsx_name = payload["files"][0]["name"] if payload["files"] else None
        check("conversion produced an .xlsx", bool(xlsx_name) and xlsx_name.endswith(".xlsx"))

        # ---- outputs --------------------------------------------------------
        r = client.get("/api/outputs")
        outputs = r.get_json()["outputs"]
        check("GET /api/outputs lists the generated workbook",
              r.status_code == 200 and any(o["name"] == xlsx_name for o in outputs))

        # ---- download -------------------------------------------------------
        r = client.get(f"/api/download/{xlsx_name}")
        check("GET /api/download streams the workbook",
              r.status_code == 200 and r.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if r.status_code == 200:
            stream = io.BytesIO(r.data)
            wb = load_workbook(stream, read_only=True)
            check("downloaded workbook has sheets", len(wb.sheetnames) > 0)
            wb.close()

        # ---- merge ----------------------------------------------------------
        r = client.post("/api/merge", json={"files": [], "mode": "sheets"})
        payload = r.get_json()
        check("POST /api/merge combines outputs",
              r.status_code == 200 and payload["ok"]
              and payload["result"]["name"].endswith(".xlsx"))

        # ---- extract --------------------------------------------------------
        r = client.post("/api/extract",
                        json={"fields": ["material_code", "quantity"]})
        payload = r.get_json()
        check("POST /api/extract consolidates columns",
              r.status_code == 200 and payload["ok"]
              and payload["result"]["rows"] >= 2
              and "Material Code" in payload["result"]["columns"])
        ext_name = payload["result"]["name"] if payload["ok"] else None
        if ext_name:
            r = client.get(f"/api/download/{ext_name}")
            stream = io.BytesIO(r.data)
            wb = load_workbook(stream, read_only=True)
            values = [str(c.value) for row in wb.worksheets[0].iter_rows()
                      for c in row if c.value is not None]
            wb.close()
            check("extracted workbook contains the row data",
                  "MC-2001" in values and "MC-2002" in values)

        # ---- extract with no fields is rejected ------------------------------
        r = client.post("/api/extract", json={"fields": []})
        check("extract with no fields is rejected", r.status_code == 400)

        # ---- encrypted needs password ----------------------------------------
        with open(sample.replace(".pdf", "_enc.pdf"), "rb") as fh:
            r = client.post("/api/upload",
                            data={"pdfs": (fh, "confidential.pdf")},
                            content_type="multipart/form-data")
        payload = r.get_json()
        check("encrypted PDF is reported as needing a password",
              payload["files"][0]["error"] == "encrypted")
        with open(sample.replace(".pdf", "_enc.pdf"), "rb") as fh:
            r = client.post("/api/upload",
                            data={"pdfs": (fh, "confidential.pdf"),
                                  "password": "secret"},
                            content_type="multipart/form-data")
        payload = r.get_json()
        check("encrypted PDF converts with the correct password",
              payload["files"][0]["ok"])

        # ---- bad file name is rejected ---------------------------------------
        r = client.get("/api/download/..%2F..%2Fetc%2Fpasswd")
        check("path traversal in download is blocked",
              r.status_code in (400, 404) and r.status_code != 200)
        r = client.get("/api/download/no_such_file.xlsx")
        check("missing download returns 404", r.status_code == 404)

        # ---- reset ------------------------------------------------------------
        r = client.post("/api/reset")
        r2 = client.get("/api/outputs")
        check("reset clears session outputs",
              r.status_code == 200 and r2.get_json()["outputs"] == [])

        # ---- magic-byte validation -------------------------------------------
        # Upload a non-PDF file with .pdf extension
        fake_pdf = os.path.join(tmp, "fake.pdf")
        with open(fake_pdf, "wb") as fh:
            fh.write(b"MZ\x90\x00" + b"\x00" * 100)  # EXE header, not PDF
        with open(fake_pdf, "rb") as fh:
            r = client.post("/api/upload",
                            data={"pdfs": (fh, "malicious.pdf")},
                            content_type="multipart/form-data")
        payload = r.get_json()
        check("magic-byte validation rejects non-PDF files",
              r.status_code == 400 and not payload["ok"])

        # ---- CSRF protection -------------------------------------------------
        # POST without X-CSRF-Token should be rejected in non-testing mode
        app_no_test = create_app(data_dir=os.path.join(tmp, "data-csrf"))
        app_no_test.config["TESTING"] = False
        client_no_test = app_no_test.test_client()
        r = client_no_test.post("/api/reset", headers={})
        check("CSRF protection rejects POSTs without token",
              r.status_code == 403)

        # ---- CSP header ------------------------------------------------------
        r = client.get("/")
        csp = r.headers.get("Content-Security-Policy", "")
        check("Content-Security-Policy header is set",
              "default-src 'self'" in csp)

        # ---- security headers ------------------------------------------------
        check("X-Content-Type-Options header is nosniff",
              r.headers.get("X-Content-Type-Options") == "nosniff")
        check("X-Frame-Options header is DENY",
              r.headers.get("X-Frame-Options") == "DENY")

        # ---- /health no longer leaks version ---------------------------------
        r = client.get("/health")
        body = r.get_json()
        check("/health endpoint does not expose version",
              "version" not in body)

        print("=" * 56)
        passed = sum(1 for _, ok in CHECKS if ok)
        print(f"Passed {passed}/{len(CHECKS)} web checks.")
        if passed != len(CHECKS):
            sys.exit(1)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == "__main__":
    main()
