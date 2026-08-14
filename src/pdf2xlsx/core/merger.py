"""Merge multiple .xlsx files into a single workbook.

Two modes are supported:

* ``sheets`` – every source sheet becomes its own worksheet in the merged
  workbook (``source – sheet``), nothing is dropped and all formatting,
  merges and images are preserved.
* ``append`` – source sheets whose schema matches (same column count and
  same header row) are stacked vertically; incompatible sheets fall back
  to separate worksheets.

openpyxl cannot read pictures back from an existing workbook, so images are
extracted from the source .xlsx packages directly (see ``_extract_images``)
and re-embedded in the merged output.
"""

import io
import os
import re
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook

from . import excelio as x
from .utils import unique_path

# OOXML namespaces
NS_SPREADSHEET = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
NS_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
NS_RELSPKG = "{http://schemas.openxmlformats.org/package/2006/relationships}"
NS_XDR = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
NS_A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"


class MergeCancelled(Exception):
    """Raised when the user cancels the running merge."""


def _normalize(name):
    """Normalised key used to group sheets across files for append mode."""
    return re.sub(r"[^0-9a-z]+", "", str(name).lower())


def _row_values(ws, r):
    return [str(c.value).strip() if c.value is not None else "" for c in ws[r]]


def _schemas_match(ws_a, ws_b):
    if ws_a.max_column != ws_b.max_column or ws_a.max_column == 0:
        return False
    return _row_values(ws_a, 1) == _row_values(ws_b, 1)


def merge_excels(input_paths, out_path, mode="sheets",
                 cancel_event=None, progress_cb=None):
    """Merge ``input_paths`` into ``out_path``. Returns ``{path, sheets, notes}``."""
    out_path = unique_path(out_path)
    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    notes = []
    image_map = {}
    for path in input_paths:
        if cancel_event is not None and cancel_event.is_set():
            raise MergeCancelled()
        image_map[path] = _extract_images(path)

    if mode == "append":
        _merge_append(input_paths, out_wb, image_map, notes, cancel_event, progress_cb)
    else:
        _merge_separate(input_paths, out_wb, image_map, notes, cancel_event, progress_cb)

    part = out_path + ".part"
    try:
        out_wb.save(part)
        os.replace(part, out_path)
    finally:
        out_wb.close()
        if os.path.exists(part):
            try:
                os.remove(part)
            except OSError:
                pass

    return {"path": out_path, "sheets": list(out_wb.sheetnames), "notes": notes}


def _open_all(input_paths, cancel_event):
    workbooks = []
    try:
        for path in input_paths:
            if cancel_event is not None and cancel_event.is_set():
                raise MergeCancelled()
            workbooks.append(load_workbook(path))
        return workbooks
    except MergeCancelled:
        for wb in workbooks:
            wb.close()
        raise


def _merge_separate(input_paths, out_wb, image_map, notes, cancel_event, progress_cb):
    from .utils import base_stem
    workbooks = _open_all(input_paths, cancel_event)
    total = len(input_paths)
    try:
        for i, wb in enumerate(workbooks):
            if cancel_event is not None and cancel_event.is_set():
                raise MergeCancelled()
            if progress_cb:
                progress_cb(i + 1, total)
            stem = base_stem(input_paths[i])
            for ws in wb.worksheets:
                title = x.prefixed_sheet_title(out_wb, stem, ws.title)
                dst = out_wb.create_sheet(title=title)
                x.copy_sheet_full(ws, dst,
                                  images=image_map[input_paths[i]].get(ws.title, []))
    finally:
        for wb in workbooks:
            wb.close()


def _merge_append(input_paths, out_wb, image_map, notes, cancel_event, progress_cb):
    from .utils import base_stem
    workbooks = _open_all(input_paths, cancel_event)
    try:
        groups = {}
        for path, wb in zip(input_paths, workbooks):
            stem = base_stem(path)
            for ws in wb.worksheets:
                groups.setdefault(_normalize(ws.title), []).append((path, stem, ws))

        processed = 0
        for entries in groups.values():
            if cancel_event is not None and cancel_event.is_set():
                raise MergeCancelled()
            path0, stem0, first_ws = entries[0]
            title = x.unique_sheet_title(out_wb, first_ws.title or stem0)
            dst = out_wb.create_sheet(title=title)
            x.copy_sheet_full(first_ws, dst, images=image_map[path0].get(first_ws.title, []))
            for path, stem, ws in entries[1:]:
                if _schemas_match(first_ws, ws):
                    appended = x.append_sheet_data(
                        ws, dst, images=image_map[path].get(ws.title, []))
                    notes.append(f"Appended {appended} row(s) into '{dst.title}' "
                                 f"from '{stem} – {ws.title}'.")
                else:
                    alt = x.prefixed_sheet_title(out_wb, stem, ws.title)
                    x.copy_sheet_full(ws, out_wb.create_sheet(title=alt),
                                      images=image_map[path].get(ws.title, []))
                    notes.append(f"'{stem} – {ws.title}' had a different schema; "
                                 f"kept as its own sheet '{alt}'.")
            processed += 1
            if progress_cb:
                progress_cb(processed, len(groups))
    finally:
        for wb in workbooks:
            wb.close()


# --------------------------------------------------------------------------
# Direct OOXML image extraction (openpyxl cannot read pictures)
# --------------------------------------------------------------------------

def _resolve_target(rels_dir, base_part, target):
    """Resolve a relationship target to a zip path inside the package."""
    if target.startswith("/"):
        return target.lstrip("/")
    if target.startswith("http") or "://" in target:
        return None
    if target.startswith("../"):
        target = os.path.normpath(os.path.join(base_part, "..", target))
        target = os.path.normpath(target)
    else:
        target = os.path.normpath(os.path.join(rels_dir, target))
    return target


def _extract_images(xlsx_path):
    """Return {sheet_title: [image records]} for every picture in the file.

    Each record is ``{data, ext, col, row, cx, cy}`` with ``col``/``row``
    being the zero-based anchor cell. Records use anchor cell + pixel size
    so they can be re-embedded through openpyxl.
    """
    result = {}
    if not os.path.exists(xlsx_path):
        return result
    try:
        zf = zipfile.ZipFile(xlsx_path)
    except zipfile.BadZipFile:
        return result

    try:
        # sheet name -> workbook-level relationship id
        wb_xml = zf.read("xl/workbook.xml")
        wb_root = ET.fromstring(wb_xml)
        sheet_rids = {}
        for sheet in wb_root.findall(f"{NS_SPREADSHEET}sheets/{NS_SPREADSHEET}sheet"):
            name = sheet.get("name")
            rid = sheet.get(f"{NS_REL}id")
            if name and rid:
                sheet_rids[name] = rid

        # workbook rels: relationship id -> worksheet path
        wb_rels = {}
        try:
            rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        except KeyError:
            rels = ET.fromstring(zf.read("xl/workbook.xml.rels"))
        for rel in rels.findall(f"{NS_RELSPKG}Relationship"):
            if "worksheet" in rel.get("Type", ""):
                wb_rels[rel.get("Id")] = _resolve_target(
                    "xl/_rels", "xl/workbook.xml", rel.get("Target"))

        for name, rid in sheet_rids.items():
            sheet_path = wb_rels.get(rid)
            if not sheet_path:
                continue
            try:
                sheet_xml = zf.read(sheet_path)
            except KeyError:
                continue
            sheet_root = ET.fromstring(sheet_xml)
            drawing_rid = None
            drawing_el = sheet_root.find(f"{NS_SPREADSHEET}drawing")
            if drawing_el is not None:
                drawing_rid = drawing_el.get(f"{NS_REL}id")
            if not drawing_rid:
                continue

            sheet_dir = os.path.dirname(sheet_path)
            rels_path = f"{sheet_dir}/_rels/{os.path.basename(sheet_path)}.rels"
            drawing_path = None
            try:
                srels = ET.fromstring(zf.read(rels_path))
                for rel in srels.findall(f"{NS_RELSPKG}Relationship"):
                    if rel.get("Id") == drawing_rid:
                        drawing_path = _resolve_target(sheet_dir, sheet_path, rel.get("Target"))
                        break
            except KeyError:
                continue
            if not drawing_path:
                continue

            try:
                draw_xml = zf.read(drawing_path)
                draw_root = ET.fromstring(draw_xml)
            except KeyError:
                continue

            drels_path = f"{os.path.dirname(drawing_path)}/_rels/{os.path.basename(drawing_path)}.rels"
            media_targets = {}
            try:
                drels = ET.fromstring(zf.read(drels_path))
                for rel in drels.findall(f"{NS_RELSPKG}Relationship"):
                    media_targets[rel.get("Id")] = _resolve_target(
                        os.path.dirname(drawing_path), drawing_path, rel.get("Target"))
            except KeyError:
                pass

            records = []
            for anchor in draw_root.iter(f"{NS_XDR}oneCellAnchor"):
                from_el = anchor.find(f"{NS_XDR}from")
                blip = anchor.find(f".//{NS_A}blip")
                if from_el is None or blip is None:
                    continue
                col_el = from_el.find(f"{NS_XDR}col")
                row_el = from_el.find(f"{NS_XDR}row")
                if col_el is None or row_el is None:
                    continue
                embed = blip.get(f"{NS_REL}embed")
                target = media_targets.get(embed)
                if not target:
                    continue
                try:
                    data = zf.read(target)
                except KeyError:
                    continue
                ext = os.path.splitext(target)[1].lstrip(".").lower() or "png"
                ext_el = anchor.find(f"{NS_XDR}ext")
                cx = cy = 0
                if ext_el is not None:
                    cx = int(float(ext_el.get("cx", 0)) / 9525)
                    cy = int(float(ext_el.get("cy", 0)) / 9525)
                records.append({"data": data, "ext": ext,
                                "col": int(col_el.text), "row": int(row_el.text),
                                "cx": cx, "cy": cy})
            if records:
                result[name] = records
    except Exception:
        pass
    finally:
        zf.close()
    return result
