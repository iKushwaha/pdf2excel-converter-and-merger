"""PDF → Excel conversion with zero data loss.

Strategy per page:
  * tables are detected with PyMuPDF and written to their own worksheet
    (``Page N – Table M``), merging split tables that continue onto the
    next page;
  * non-table text blocks and embedded images are written to a
    ``Page N – Text & Images`` worksheet in reading order;
  * the source PDF is processed page-by-page so large documents stay
    memory-friendly, and the result is written to a temp file that is only
    renamed into place once the whole conversion succeeded.
"""

import os
import tempfile
import threading

from openpyxl import Workbook
from openpyxl.styles import Font

try:
    import pymupdf as fitz  # preferred module name (1.24+)
except ImportError:          # pragma: no cover - older releases
    import fitz              # noqa: F401

from . import excelio as x
from .utils import unique_path

CONTINUATION_MARGIN = 60.0   # points: a table that ends this close to the
                             # bottom edge of a page may continue next page
X_TOLERANCE = 14.0           # points: column-alignment tolerance when matching
MAX_IMG_WIDTH_PX = 620.0     # embed images at most this wide (pixels)
WARNING_HEX = "E05656"       # color used for OCR/scan notices


class ConversionError(Exception):
    """Base error raised by the converter."""


class EncryptedPDF(ConversionError):
    """Raised when the PDF is encrypted and no (correct) password is known."""


class ConversionCancelled(ConversionError):
    """Raised when the user cancels the running conversion."""


class _RunState:
    """Tracks table-run identity so tables split over pages are merged."""

    def __init__(self):
        self.counter = 0
        self.last = None  # dict describing the previous table for continuation


def convert_pdf_to_excel(pdf_path, out_dir, password=None,
                         cancel_event=None, progress_cb=None):
    """Convert one PDF into a .xlsx in ``out_dir``.

    Returns a dict: ``{path, sheets, warnings}``.
    Raises ``EncryptedPDF``, ``ConversionCancelled`` or ``ConversionError``.
    """
    out_path = unique_path(os.path.join(out_dir,
                                        os.path.splitext(os.path.basename(pdf_path))[0] + ".xlsx"))
    doc = _open_document(pdf_path, password)

    if doc.page_count == 0:
        doc.close()
        raise ConversionError("The PDF contains no pages.")

    wb = Workbook()
    wb.remove(wb.active)
    state = _RunState()
    warnings = []

    try:
        for pno in range(doc.page_count):
            if cancel_event is not None and cancel_event.is_set():
                raise ConversionCancelled()
            if progress_cb:
                progress_cb(pno, doc.page_count,
                            f"Page {pno + 1} of {doc.page_count}")
            _process_page(wb, doc, doc[pno], pno, state, warnings)
    except ConversionCancelled:
        wb.close()
        doc.close()
        raise
    except Exception as exc:  # surface low-level extraction failures cleanly
        wb.close()
        doc.close()
        raise ConversionError(f"Extraction failed: {exc}") from exc

    part = out_path + ".part"
    try:
        wb.save(part)
        os.replace(part, out_path)
    finally:
        wb.close()
        doc.close()
        if os.path.exists(part):
            try:
                os.remove(part)
            except OSError:
                pass

    return {"path": out_path, "sheets": list(wb.sheetnames), "warnings": warnings}


def _open_document(pdf_path, password):
    """Open a PDF handling encryption. Raises EncryptedPDF when needed."""
    try:
        doc = fitz.open(pdf_path)
    except Exception as exc:
        raise ConversionError(f"Cannot open PDF – file is corrupt or not a PDF: {exc}") from exc
    if doc.is_encrypted:
        if not password:
            doc.close()
            raise EncryptedPDF(pdf_path)
        if not doc.authenticate(password):
            doc.close()
            raise ConversionError("Incorrect password – the PDF could not be unlocked.")
    return doc


def _process_page(wb, doc, page, pno, state, warnings):
    """Extract and write everything present on one PDF page."""
    page_height = page.rect.height
    tables = _find_tables(page)
    images = page.get_image_info(xrefs=True)
    text_blocks = [b for b in page.get_text("blocks") if b[6] == 0]

    text_sheet = wb.create_sheet(title=x.unique_sheet_title(
        wb, f"Page {pno + 1} – Text & Images"))
    title_cell = text_sheet.cell(row=1, column=1, value=f"Page {pno + 1}")
    title_cell.font = Font(bold=True, size=13)
    text_sheet.column_dimensions["A"].width = 110
    row = 2

    # One ordered stream of events: text blocks, images and tables by position.
    events = []
    for blk in text_blocks:
        events.append(("text", blk, blk[1]))
    for t in tables:
        events.append(("table", t, t.bbox[1]))
    for img in images:
        events.append(("image", img, img["bbox"][1]))
    events.sort(key=lambda e: (e[2], e[0]))

    for kind, obj, _y in events:
        if kind == "table":
            row = _write_table(wb, page, pno, obj, state, text_sheet, row,
                               page_height, warnings)
        elif kind == "text":
            if _inside_any_table(obj, tables):
                continue  # this text belongs to a table, already captured
            cell = text_sheet.cell(row=row, column=1, value=obj[4])
            cell.font = Font(size=10)
            cell.alignment = x.Alignment(wrap_text=True, vertical="top")
            row += 1
        elif kind == "image":
            if _inside_any_table(obj, tables):
                continue
            row = _embed_image(doc, text_sheet, row, obj, warnings)

    if not text_blocks and not tables and images:
        note = ("[This page appears to be a scanned image – the picture was "
                "embedded as-is. No OCR is bundled, so the page has no "
                "searchable text. Run OCR on the PDF first if you need it.]")
        cell = text_sheet.cell(row=row, column=1, value=note)
        cell.font = Font(italic=True, size=9, color=WARNING_HEX)
        row += 1
        warnings.append(f"Page {pno + 1} contains only images (scanned page); "
                        "no text extracted – OCR not enabled.")

    x.autofit_row_heights(text_sheet)
    return text_sheet


def _find_tables(page):
    """Detect tables with the ruled-line strategy.

    PDFs whose tables have no ruled lines at all (pure whitespace-separated
    columns) are not auto-detected — their text is kept intact on the text
    worksheet instead, so nothing is lost. (Heuristic "text alignment" table
    detection was rejected after testing because it mangled normal prose into
    false tables.)
    """
    try:
        return page.find_tables().tables
    except Exception:
        return []


def _bbox_of(obj):
    """Return (x0, y0, x1, y1) for a text-block tuple or image-info dict."""
    if isinstance(obj, dict):
        if "bbox" in obj:
            return obj["bbox"]
        if "x0" in obj:
            return (obj["x0"], obj["y0"], obj["x1"], obj["y1"])
    if isinstance(obj, (tuple, list)) and len(obj) >= 4:
        return (obj[0], obj[1], obj[2], obj[3])
    return None


def _inside_any_table(obj, tables):
    """True if an object's center point falls inside any detected table."""
    bbox = _bbox_of(obj)
    if bbox is None:
        return False
    cx = (bbox[0] + bbox[2]) / 2.0
    cy = (bbox[1] + bbox[3]) / 2.0
    for t in tables:
        x0, y0, x1, y1 = t.bbox
        if x0 <= cx <= x1 and y0 <= cy <= y1:
            return True
    return False


def _write_table(wb, page, pno, table, state, text_sheet, row, page_height, warnings):
    """Write a detected table to its own worksheet (or continue a split one)."""
    try:
        data = table.extract()
    except Exception as exc:
        warnings.append(f"Table on page {pno + 1} could not be read: {exc}")
        return row
    data = _clean_table_data(data)
    if not data:
        return row

    x0, y0, x1, y1 = table.bbox
    ended_at_bottom = y1 >= page_height - CONTINUATION_MARGIN

    prev = state.last
    if prev is not None and _should_continue(prev, table, x0, x1):
        ws = prev["sheet"]
        if data[0] == prev["header"]:
            data = data[1:]  # repeated header on the continuation page
        start = ws.max_row + 1
        for r, values in enumerate(data):
            for c, value in enumerate(values):
                x.write_cell(ws, start + r, c + 1, value)
        x.autofit_columns(ws)
        state.last = {"sheet": ws, "header": prev["header"], "x0": x0, "x1": x1,
                      "cols": table.col_count, "ended_at_bottom": ended_at_bottom}
    else:
        state.counter += 1
        run_id = state.counter
        title = x.unique_sheet_title(wb, f"Page {pno + 1} – Table {run_id}")
        ws = wb.create_sheet(title=title)
        for c, value in enumerate(data[0]):
            x.write_cell(ws, 1, c + 1, value, header=True)
        for r, values in enumerate(data[1:], start=2):
            for c, value in enumerate(values):
                x.write_cell(ws, r, c + 1, value)
        ws.freeze_panes = "A2"
        x.autofit_columns(ws)
        state.last = {"sheet": ws, "header": data[0], "x0": x0, "x1": x1,
                      "cols": table.col_count, "ended_at_bottom": ended_at_bottom}
        note = text_sheet.cell(row=row, column=1,
                               value=f"→ Table {run_id} is on the "
                                     f"'{title}' worksheet.")
        note.font = Font(italic=True, size=9, color="4C9AFF")
        row += 1
    return row


def _clean_table_data(data):
    """Drop fully-empty rows/columns and None-padded tails."""
    cleaned = []
    for r_values in data:
        r_values = [v if v is not None else "" for v in r_values]
        if all(not str(v).strip() for v in r_values):
            continue
        cleaned.append(r_values)
    if not cleaned:
        return []
    cols = max(len(r) for r in cleaned)
    cleaned = [r + [""] * (cols - len(r)) for r in cleaned]
    return cleaned


def _should_continue(prev, table, x0, x1):
    """Heuristic: is ``table`` the continuation of the previous page's table?"""
    if not prev.get("ended_at_bottom"):
        return False
    if prev["cols"] != table.col_count:
        return False
    if abs(prev["x0"] - x0) > X_TOLERANCE or abs(prev["x1"] - x1) > X_TOLERANCE:
        return False
    return True


def _embed_image(doc, text_sheet, row, img_info, warnings):
    """Embed an extracted PDF image into the text worksheet."""
    xref = img_info.get("xref")
    if not xref:
        return row
    try:
        info = doc.extract_image(xref)
        blob = info["image"]
        ext = info["ext"] or "png"
        fd, tmp_path = tempfile.mkstemp(suffix=f".{ext}")
        with os.fdopen(fd, "wb") as fh:
            fh.write(blob)

        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(tmp_path)
        img_w, img_h = info.get("width", 0), info.get("height", 0)
        if img_w and img_h:
            scale = min(1.0, MAX_IMG_WIDTH_PX / img_w)
            img.width = int(img_w * scale)
            img.height = int(img_h * scale)
        img.anchor = f"A{row}"
        text_sheet.add_image(img)

        caption = text_sheet.cell(row=row, column=1, value=f"[Image embedded – {info.get('width', '?')}x{info.get('height', '?')} px]")
        caption.font = Font(italic=True, size=9, color="888888")
        row += 1
    except Exception as exc:
        warnings.append(f"Image on page could not be embedded: {exc}")
    return row
